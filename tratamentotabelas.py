import os
import pandas as pd
import shutil
from datetime import datetime
from openpyxl import load_workbook


# Lista de subdiretórios a serem processados
SUBDIRETORIOS = [
    "DWNLD\\LB",
    "DWNLD\\LPC",
    "DWNLD\\TaJT",
    "DWNLD\\TaMDV",
    "DWNLD\\TaTR",
    "DWNLD\\TaVS",
    "DWNLD\\VGD",
    "DWNLD\\VGPC",
    "DWNLD\\VGPC2"
]

# Mapeamento para renomear as abas
MAPEAMENTO_ABAS = {
    "LB": "LacunasBacklinks",
    "LPC": "LacunasPalavrasChave",
    "TaJT": "JornadaTrafego",
    "TaMDV": "MediaDuracaoVisita",
    "TaTR": "TaxaRejeicao",
    "TaVS": "VisitasSite",
    "VGD": "VisaoGeralDominio",
    "VGPC": "VisaoGeralPalavrasChave",
    "VGPC2": "VisaoGeralPalavrasChave2"
}

# Dicionário para mapear meses em string para números
MES_MAP = {
    "jan": "01", "fev": "02", "mar": "03", "abr": "04", 
    "mai": "05", "jun": "06", "jul": "07", "ago": "08", 
    "set": "09", "out": "10", "nov": "11", "dez": "12"
}

# Mapeamento das categorias
CATEGORIAS = {
    "Times de Futebol": ["FLAMENGO", "CORINTHIANS", "BAYERN DE MUNIQUE", "FUTEBOL", "PALMEIRAS", "VASCO", "BOTAFOGO", "FLUMINENSE", "CRUZEIRO", "SANTOS", "INTERNACIONAL", "SÃO PAULO", "ARSENAL", "LIVERPOOL", "GRÊMIO", "ATLÉTICO", "RB BRAGANTINO", "BARCELONA", "BRASIL", "MANCHESTER CITY", "PARIS SAINT GERMAIN", "NFL"],
    "Personagens, Jogos e Desenhos": ["GALINHA PINTADINHA", "BARBIE", "PLAYSTATION", "FIFA", "TURMA DO NEYMAR JR.", "FISHER PRICE", "LOL", "URSINHOS CARINHOSOS", "HOT WHEELS", "CAPITAO AMERICA","REI LEAO","UNICORNIO", "PETS", "ONE PIECE", "MINECRAFT", "ROBLOX", "HOMEM ARANHA", "HARRY POTTER", "PATRULHA CANINA", "FREE FIRE", "MARIA CLARA E JP", "SONIC", "STITCH", "NARUTO", "BOB ESPONJA", "CARROS", "PEPPA PIG", "POKEMON", "STRANGER THINGS", "AMONG US", "BOLOFOFOS", "DINOSSAURO", "MOANA", "BATMAN", "MARIO", "MUNDO BITA", "WANDINHA", "BRANCA DE NEVE", "ENALDINHO", "FLASH", "LULUCA", "BABY ALIVE", "BABY SHARK", "FROZEN", "MULHER MARAVILHA", "RICK AND MORTY", "TOY STORY", "TURMA DA MONICA", "CINDERELA", "MINIONS", "MY LITTLE PONY", "PANTERA NEGRA", "POCOYO", "DORA AVENTUREIRA", "MICKEY", "MONICA", "ARIEL", "BLUEY", "MINNIE", "MORANGUINHO", "VINGADORES", "LUCAS NETO", "PRINCESAS", "UNICÓRNIO", "COCOMELON", "POLLY POCKET", "TARTARUGAS NINJAS", "JASMINE", "PKXD", "PRINCESA SOFIA", "REI LEÃO", "CAPITÃO AMÉRICA", "SUPER MAN", "REBECCA BONBON", "AUTHENTIC GAMES", "MARIE", "PLAY DOH", "PJ MASK", "SPIDEY", "MR POTATO HEAD", "GATO GALÁCTICO", "POP FUN", "SUPER MARIO", "SIMPSONS", "LIGA DA JUSTIÇA", "MENINAS SUPERPODEROSAS", "TROLLS", "ONDE ESTÁ WALLY", "DC SUPER HERO GIRLS", "A CASA MÁGICA DA GABY", "NEYMAR JR.", "TURMA DA MATA", "DC SUPER FRIENDS", "DC ORIGINALS", "JURASSIC WORLD", "SHREK", "PAC MAN", "MTV", "ONDE ESTÁ WALLY", "NEYMAR JR."]
}

def determinar_categoria(valor):
    for categoria, itens in CATEGORIAS.items():
        if valor in itens:
            return categoria
    return "Desconhecido"

def apagar_arquivos(diretorios):
    """Apaga todos os arquivos dentro dos subdiretórios especificados."""
    for subdiretorio in diretorios:
        caminho_diretorio = os.path.join(os.path.abspath(os.path.dirname(__file__)), subdiretorio)
        if os.path.exists(caminho_diretorio):
            for filename in os.listdir(caminho_diretorio):
                file_path = os.path.join(caminho_diretorio, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                        print(f"Arquivo {file_path} apagado.")
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                        print(f"Diretório {file_path} apagado.")
                except Exception as e:
                    print(f"Erro ao apagar {file_path}: {e}")
        else:
            print(f"Diretório não encontrado: {caminho_diretorio}")

def renomear_arquivos_csv(diretorios):
    """Renomeia os arquivos CSV para o mesmo nome da pasta em que estão."""
    for subdiretorio in diretorios:
        caminho_diretorio = os.path.join(os.path.abspath(os.path.dirname(__file__)), subdiretorio)
        
        if os.path.exists(caminho_diretorio):
            for filename in os.listdir(caminho_diretorio):
                if filename.endswith('.csv'):
                    caminho_antigo = os.path.join(caminho_diretorio, filename)
                    caminho_novo = os.path.join(caminho_diretorio, f"{os.path.basename(caminho_diretorio)}.csv")
                    
                    if os.path.isfile(caminho_antigo) and not os.path.exists(caminho_novo):
                        try:
                            os.rename(caminho_antigo, caminho_novo)
                            print(f"Renomeado {caminho_antigo} para {caminho_novo}")
                        except Exception as e:
                            print(f"Erro ao renomear {caminho_antigo} para {caminho_novo}: {e}")
                    else:
                        print(f"Arquivo {caminho_novo} já existe. Pulando renomeação.")
        else:
            print(f"Diretório não encontrado: {caminho_diretorio}")

def converter_csv_para_xlsx(diretorios, arquivo_saida):
    """Converte todos os arquivos CSV em cada subdiretório para um único arquivo Excel com múltiplas abas."""
    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
        for subdiretorio in diretorios:
            caminho_diretorio = os.path.join(os.path.abspath(os.path.dirname(__file__)), subdiretorio)
            
            if os.path.exists(caminho_diretorio):
                for filename in os.listdir(caminho_diretorio):
                    if filename.endswith('.csv'):
                        caminho_csv = os.path.join(caminho_diretorio, filename)
                        
                        if os.path.isfile(caminho_csv):
                            nome_aba = os.path.basename(caminho_diretorio)  # Nome da aba será o nome do subdiretório
                            
                            try:
                                df = pd.read_csv(caminho_csv)
                                df.to_excel(writer, sheet_name=nome_aba, index=False)
                                print(f"Adicionando dados à aba {nome_aba}...")
                            except Exception as e:
                                print(f"Erro ao ler {caminho_csv}: {e}")
                        else:
                            print(f"Arquivo não encontrado: {caminho_csv}")
            else:
                print(f"Diretório não encontrado: {caminho_diretorio}")

    print(f"Arquivo Excel salvo em: {arquivo_saida}")


def renomear_abas(arquivo_saida):
    """Renomeia as abas do arquivo Excel de acordo com o mapeamento fornecido."""
    print("Renomeando abas do arquivo Excel...")
    with pd.ExcelWriter(arquivo_saida, engine='openpyxl', mode='a') as writer:
        workbook = writer.book
        for aba in workbook.sheetnames:
            if aba in MAPEAMENTO_ABAS:
                novo_nome = MAPEAMENTO_ABAS[aba]
                workbook[aba].title = novo_nome
                print(f"Aba '{aba}' renomeada para '{novo_nome}'")

def remover_sufixo_dos_dominios(arquivo_saida):
    """Remove o sufixo '.com.br' do cabeçalho das colunas nas abas especificadas."""
    print("Removendo sufixo '.com.br' dos cabeçalhos das colunas...")
    abas_para_processar = [
        "LacunasBacklinks",
        "LacunasPalavrasChave",
        "MediaDuracaoVisita",
        "TaxaRejeicao",
        "VisitasSite"
    ]
    
    with pd.ExcelWriter(arquivo_saida, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        for nome_aba in abas_para_processar:
            try:
                df = pd.read_excel(arquivo_saida, sheet_name=nome_aba)
                if 'Unnamed: 0' in df.columns:
                    df.rename(columns={'Unnamed: 0': 'Mes'}, inplace=True)
                df.columns = [col.replace('.com.br', '') for col in df.columns]
                df.to_excel(writer, sheet_name=nome_aba, index=False)
                print(f"Sufixo '.com.br' removido da aba '{nome_aba}'")
            except Exception as e:
                print(f"Erro ao processar a aba '{nome_aba}': {e}")

def converter_tempo_para_segundos(arquivo_saida):
    """Converte o tempo médio de hh:mm para segundos nas abas especificadas."""
    print("Convertendo tempo médio de hh:mm para segundos nas abas especificadas...")
    aba_para_processar = "MediaDuracaoVisita"
    
    try:
        df = pd.read_excel(arquivo_saida, sheet_name=aba_para_processar)
        if 'Unnamed: 0' in df.columns:
            df.rename(columns={'Unnamed: 0': 'Mes'}, inplace=True)
        for col in df.columns[1:]:  # Ignorar a primeira coluna
            df[col] = df[col].apply(lambda x: sum(int(i) * 60**index for index, i in enumerate(reversed(str(x).split(':')))))
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=aba_para_processar, index=False)
        print(f"Tempo convertido para segundos na aba '{aba_para_processar}'")
    except Exception as e:
        print(f"Erro ao processar a aba '{aba_para_processar}': {e}")

def converter_para_numero(arquivo_saida):
    """Converte valores para número na aba 'TaxaRejeicao'."""
    print("Convertendo valores para número na aba 'TaxaRejeicao'...")
    aba_para_processar = "TaxaRejeicao"
    
    try:
        df = pd.read_excel(arquivo_saida, sheet_name=aba_para_processar)
        if 'Unnamed: 0' in df.columns:
            df.rename(columns={'Unnamed: 0': 'Mes'}, inplace=True)
        for col in df.columns[1:]:  # Ignorar a primeira coluna
            df[col] = df[col].apply(lambda x: float(str(x).replace('%', '').replace(',', '.')) if isinstance(x, (int, float, str)) else x)
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=aba_para_processar, index=False)
        print(f"Valores convertidos para número na aba '{aba_para_processar}'")
    except Exception as e:
        print(f"Erro ao processar a aba '{aba_para_processar}': {e}")


def merge_abas_openpyxl(arquivo_saida, aba1_nome, aba2_nome):
    """Faz o merge das abas especificadas no arquivo Excel utilizando openpyxl."""
    print(f"Realizando o merge das abas '{aba1_nome}' e '{aba2_nome}' utilizando openpyxl...")
    try:
        # Carregar o workbook e as duas abas
        workbook = load_workbook(arquivo_saida)
        aba1 = workbook[aba1_nome]
        aba2 = workbook[aba2_nome]

        # Encontrar a última linha preenchida na aba1
        last_row = aba1.max_row

        # Copiar os dados da aba2 para aba1
        for row in aba2.iter_rows(min_row=2, values_only=True):  # Começar na segunda linha para ignorar o cabeçalho
            aba1.append(row)

        # Salvar as alterações no arquivo
        workbook.save(arquivo_saida)
        print(f"Merge das abas '{aba1_nome}' e '{aba2_nome}' concluído com sucesso utilizando openpyxl.")
    except Exception as e:
        print(f"Erro ao realizar o merge das abas '{aba1_nome}' e '{aba2_nome}': {e}")


def apagar_aba(arquivo_saida, aba_nome):
    """Apaga a aba especificada do arquivo Excel."""
    print(f"Apagando a aba '{aba_nome}'...")
    try:
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl', mode='a') as writer:
            workbook = writer.book
            if aba_nome in workbook.sheetnames:
                std = workbook[aba_nome]
                workbook.remove(std)
                print(f"Aba '{aba_nome}' apagada com sucesso.")
            else:
                print(f"Aba '{aba_nome}' não encontrada.")
    except Exception as e:
        print(f"Erro ao apagar a aba '{aba_nome}': {e}")

def categorizar_abas(arquivo_saida):
    """Adiciona a categoria para cada valor na aba 'VisaoGeralPalavrasChave'."""
    print(f"Adicionando categorias à aba 'VisaoGeralPalavrasChave'...")
    aba_para_processar = "VisaoGeralPalavrasChave"
    
    try:
        df = pd.read_excel(arquivo_saida, sheet_name=aba_para_processar)
        if 'Unnamed: 0' in df.columns:
            df.rename(columns={'Unnamed: 0': 'Mes'}, inplace=True)
        df['Categoria'] = df.iloc[:, 0].apply(determinar_categoria)  # Aplica a função determinar_categoria na primeira coluna
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=aba_para_processar, index=False)
        print(f"Categorias adicionadas à aba '{aba_para_processar}'")
    except Exception as e:
        print(f"Erro ao adicionar categorias à aba '{aba_para_processar}': {e}")
        
def categorizar_palavras_lpc(arquivo_saida):
    """Categorização de palavras na aba 'LacunasPalavrasChave', com base na primeira palavra da Coluna A."""
    print("Categorizando palavras na aba 'LacunasPalavrasChave'...")

    # Definindo as categorias e palavras iniciais correspondentes
    categorias = {
        "Mala": "MALA",
        "Mochila": "MOCHILA",
        "Bolsa": "BOLSA",
        "Carteira": "CARTEIRA",
        "Garrafa": "GARRAFA",
        "Fresqueira": "FRESQUEIRA",
        "Lancheira": "LANCHEIRA"
    }

    try:
        # Carregando a aba "LacunasPalavrasChave"
        df = pd.read_excel(arquivo_saida, sheet_name="LacunasPalavrasChave")
        
        # Renomeando a coluna 'Unnamed: 0' para 'Mes' se existir
        if 'Unnamed: 0' in df.columns:
            df.rename(columns={'Unnamed: 0': 'Mes'}, inplace=True)
        
        # Criando uma nova coluna "Categoria" (Coluna R)
        def atribuir_categoria(valor):
            primeira_palavra = valor.split()[0].upper()  # Pega a primeira palavra em maiúscula
            for categoria, palavra in categorias.items():
                if primeira_palavra.startswith(palavra):
                    return categoria
            return "Outros"

        # Aplicando a função em todas as linhas da Coluna A
        df["Categoria"] = df.iloc[:, 0].apply(atribuir_categoria)

        # Salvando as alterações no arquivo Excel
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name="LacunasPalavrasChave", index=False)
        
        print("Categorização concluída na aba 'LacunasPalavrasChave'.")
    except Exception as e:
        print(f"Erro ao categorizar palavras na aba 'LacunasPalavrasChave': {e}")


def converter_mes_para_data(mes_str):
    """Converte o mês de formato 'jan. de 2024' para '2024-01-01' como datetime."""
    try:
        mes_abrev, ano = mes_str.split(' de ')
        mes_num = MES_MAP[mes_abrev.lower()[:3]]
        return datetime.strptime(f"01/{mes_num}/{ano}", "%d/%m/%Y")
    except Exception as e:
        print(f"Erro ao converter '{mes_str}': {e}")
        return None  # Retorna None em caso de erro

def converter_coluna_mes_para_data(arquivo_saida, aba_nome):
    """Aplica a conversão da coluna 'Mes' para formato de data nas abas especificadas."""
    print(f"Convertendo coluna 'Mes' para formato 'YYYY-MM-DD' na aba '{aba_nome}'...")
    try:
        df = pd.read_excel(arquivo_saida, sheet_name=aba_nome)
        if 'Mes' in df.columns:
            df['Mes'] = df['Mes'].apply(converter_mes_para_data)
            with pd.ExcelWriter(arquivo_saida, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=aba_nome, index=False)
            print(f"Conversão para data concluída na aba '{aba_nome}'.")
        else:
            print(f"A coluna 'Mes' não foi encontrada na aba '{aba_nome}'.")
    except Exception as e:
        print(f"Erro ao processar a aba '{aba_nome}': {e}")


def transpor_jornada_trafego(arquivo_saida):
    """Transpõe a aba 'JornadaTrafego' para inverter linhas e colunas."""
    print("Transpondo a aba 'JornadaTrafego'...")
    
    try:
        df = pd.read_excel(arquivo_saida, sheet_name='JornadaTrafego')
        
        # Transpor a tabela
        df_transposta = df.set_index('Destino').T.reset_index()
        df_transposta.rename(columns={'index': 'Canal'}, inplace=True)
        
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_transposta.to_excel(writer, sheet_name='JornadaTrafego', index=False)
        
        print("Aba 'JornadaTrafego' transposta com sucesso.")
    except Exception as e:
        print(f"Erro ao transpor a aba 'JornadaTrafego': {e}")


def transformar_visao_geral_dominio(arquivo_saida):
    """Transforma a aba 'VisaoGeralDominio' convertendo colunas de datas em linhas, aplicando a formatação de data e somando os valores das colunas."""
    print("Transformando a aba 'VisaoGeralDominio'...")

    try:
        # Carregando a aba "VisaoGeralDominio"
        df = pd.read_excel(arquivo_saida, sheet_name="VisaoGeralDominio")
        
        # Remover a coluna 'Summary' (Coluna E)
        df.drop(columns=['Summary'], inplace=True)
        
        # Derreter as colunas de data para transformar em linhas
        df_melted = pd.melt(df, id_vars=['Target', 'Target Type', 'Metric', 'Database'], var_name='Data', value_name='Valor')
        
        # Converter as colunas da métrica em colunas individuais e a coluna de data em formato de data
        df_pivot = df_melted.pivot_table(index=['Target', 'Data'], columns='Metric', values='Valor', aggfunc='first').reset_index()
        
        # Convertendo a coluna 'Data' para o formato de data 'YYYY-MM-DD'
        df_pivot['Data'] = pd.to_datetime(df_pivot['Data'] + '-01', format='%Y-%m-%d')
        
        # Adicionar uma nova coluna com a soma de todas as colunas de métricas
        df_pivot['Soma_Total'] = df_pivot.iloc[:, 2:].sum(axis=1)  # Soma todas as colunas de métricas (exclui 'Target' e 'Data')

        # Salvando as alterações no arquivo Excel
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_pivot.to_excel(writer, sheet_name="VisaoGeralDominio", index=False)
        
        print("Transformação concluída na aba 'VisaoGeralDominio'.")
    except Exception as e:
        print(f"Erro ao transformar a aba 'VisaoGeralDominio': {e}")


def preencher_celulas_vazias_com_zero(arquivo_saida):
    """Percorre todas as abas de um arquivo Excel e substitui células vazias por '0'."""
    print(f"Preenchendo células vazias com '0' em todas as abas do arquivo '{arquivo_saida}'...")
    
    try:
        # Carregar o workbook
        workbook = load_workbook(arquivo_saida)
        
        # Percorrer todas as abas
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"Processando a aba '{sheet_name}'...")
            
            # Percorrer todas as células da aba
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None or cell.value == "":
                        cell.value = "0"

        # Salvar as alterações no arquivo
        workbook.save(arquivo_saida)
        print("Preenchimento de células vazias concluído com sucesso!")
    
    except Exception as e:
        print(f"Erro ao preencher células vazias: {e}")


def adicionar_data_extracao(arquivo_saida):
    """Adiciona a coluna 'Data_Extração' com a data atual em todas as abas do arquivo Excel sem recriar abas."""
    data_extracao = datetime.now().strftime('%Y-%m-%d')
    print(f"Adicionando coluna 'Data_Extracao' com a data: {data_extracao}")

    try:
        # Carregar o arquivo Excel existente
        workbook = load_workbook(arquivo_saida)

        # Iterar sobre todas as abas do arquivo
        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]

            # Verificar se a coluna 'Data_Extração' já existe
            if 'Data_Extracao' not in [cell.value for cell in sheet[1]]:
                # Adicionar a coluna 'Data_Extração' em todas as linhas
                col_index = sheet.max_column + 1  # Próxima coluna disponível
                sheet.cell(row=1, column=col_index).value = 'Data_Extracao'
                
                for row in range(2, sheet.max_row + 1):
                    # Adiciona a data no formato datetime para garantir compatibilidade com SQL
                    sheet.cell(row=row, column=col_index).value = datetime.now()

        # Salvar o arquivo com as mudanças
        workbook.save(arquivo_saida)
        print("Coluna 'Data_Extracao' adicionada com sucesso a todas as abas.")

    except Exception as e:
        print(f"Erro ao adicionar a coluna 'Data_Extracao': {e}")


def main():
    caminho_atual = os.path.abspath(os.path.dirname(__file__))
    diretorio_saida = os.path.join(caminho_atual, "ExcelTratado")
    if not os.path.exists(diretorio_saida):
        os.makedirs(diretorio_saida)
    arquivo_saida = os.path.join(diretorio_saida, "DadosTratados.xlsx")
    
    renomear_arquivos_csv(SUBDIRETORIOS)
    converter_csv_para_xlsx(SUBDIRETORIOS, arquivo_saida)
    renomear_abas(arquivo_saida)
    remover_sufixo_dos_dominios(arquivo_saida)
    converter_tempo_para_segundos(arquivo_saida)
    converter_para_numero(arquivo_saida)
    merge_abas_openpyxl(arquivo_saida, "VisaoGeralPalavrasChave", "VisaoGeralPalavrasChave2")
    categorizar_abas(arquivo_saida)
    transpor_jornada_trafego(arquivo_saida)
    categorizar_palavras_lpc(arquivo_saida)
    transformar_visao_geral_dominio(arquivo_saida)
    apagar_aba(arquivo_saida, "VisaoGeralPalavrasChave2")
    preencher_celulas_vazias_com_zero(arquivo_saida)
    adicionar_data_extracao(arquivo_saida)
    apagar_arquivos(SUBDIRETORIOS)
    
    # Aplicar a conversão da coluna 'Mes' para data nas abas especificadas
    converter_coluna_mes_para_data(arquivo_saida, "MediaDuracaoVisita")
    converter_coluna_mes_para_data(arquivo_saida, "TaxaRejeicao")
    converter_coluna_mes_para_data(arquivo_saida, "VisitasSite")
    
    print("Conversão concluída!")

if __name__ == "__main__":
    main()
